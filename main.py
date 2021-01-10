# !/usr/bin/env python
# -*- coding: utf-8 -*-
# Author: LiangKaimeng
# Date  : 2021/1/10

import os
import warnings
import pandas as pd
from openpyxl import load_workbook
from AutoExcelMake.excel_work_path import extract_file_name as efn

warnings.filterwarnings('ignore')
# 初始化目录
root_path = os.getcwd()


def load_excelbook(path):
    return load_workbook(path)


def load_data(path):
    data = pd.read_excel(path)
    data['团长uid'] = data['团长uid'].astype(str).map(lambda x: "'" + x)
    
    print(data.values)





    # ws = wb['团长分层']
    # ws.cell(row=4, column=4).value = 1
    # wb.save(os.getcwd() + '\工作表' + '\\' + os.path.splitext(file_name)[0] + time.strftime("%Y%m%d",
    #                                                                                      time.localtime()) + '.xlsx')
    # wb.close()



if __name__ == '__main__':
    # path = efn(root_path+'\备份表\\', 'test')
    # load_excelbook(path)
    load_data(root_path + '\明细表\\' + '团长分层明细.xlsx')